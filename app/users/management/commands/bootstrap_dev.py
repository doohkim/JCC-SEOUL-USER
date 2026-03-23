"""
로컬 초기 세팅 한 번에:

1. 슈퍼유저 ``admin`` / ``1234`` (bootstrap_admin)
2. 조직도 (seed_org_chart)
3. 엑셀 → 청년부 팀·교적 Member·소속 (seed_youth_roster, 시트 기본 ``주일 88``)
4. 같은 엑셀에서 주일 예배 출석 시트 자동 탐지 후 ``SundayAttendanceLine`` 저장
   (``--sunday-sheet`` 로 수동 지정 가능; 출석은 부서+예배일 단위, 주차 테이블 없음)

엑셀 기본 경로: ``~/Downloads/2026 예배 출석 명단.xlsx`` (다르면 ``--xlsx``)

사용::

    python manage.py bootstrap_dev
    python manage.py bootstrap_dev --xlsx /path/to/명단.xlsx
    python manage.py bootstrap_dev --xlsx ~/Downloads/book.xlsx --sunday-sheet "26.03.22 주일예배"
    python manage.py bootstrap_dev --skip-excel   # 엑셀 없이 admin+조직만
"""

from __future__ import annotations

import os
import re
from datetime import date
from pathlib import Path

from django.contrib.auth import get_user_model
from django.core.management import call_command
from django.core.management.base import BaseCommand, CommandError

from users.models import RoleLevel

DEFAULT_XLSX = Path.home() / "Downloads" / "2026 예배 출석 명단.xlsx"
WORKBOOK_BASENAME = "2026 예배 출석 명단.xlsx"


def _app_dir() -> Path:
    return Path(__file__).resolve().parents[3]


def resolve_attendance_workbook(cli_path_str: str) -> Path | None:
    """
    엑셀 탐색 순서: ``--xlsx`` 값 → ``JCC_SEED_XLSX`` → ``app/fixtures/`` → ``~/Downloads/``.
    """
    candidates: list[Path] = []
    candidates.append(Path(cli_path_str).expanduser().resolve())
    env = (os.environ.get("JCC_SEED_XLSX") or "").strip()
    if env:
        candidates.append(Path(env).expanduser().resolve())
    candidates.append(_app_dir() / "fixtures" / WORKBOOK_BASENAME)
    candidates.append(Path.home() / "Downloads" / WORKBOOK_BASENAME)
    seen: set[str] = set()
    for p in candidates:
        key = str(p.resolve())
        if key in seen:
            continue
        seen.add(key)
        if p.is_file():
            return p.resolve()
    return None


def _sheet_title_date(name: str) -> date | None:
    """시트명 앞쪽 ``YY.MM.DD`` → 날짜 (예: ``26.03.22 주일예배``)."""
    m = re.search(r"(\d{2})\.(\d{1,2})\.(\d{1,2})", name)
    if not m:
        return None
    yy, mo, d = int(m.group(1)), int(m.group(2)), int(m.group(3))
    year = 2000 + yy if yy < 100 else yy
    try:
        return date(year, mo, d)
    except ValueError:
        return None


def _pick_sunday_attendance_sheet(path: Path) -> str | None:
    """
    ``주일`` + ``예배`` + 날짜가 있는 시트 중, **시트명 날짜가 가장 최근**인 것을 고름.
    ``주일 88``·``주일 인천`` 등 로스터/집회 시트는 제외.
    """
    import openpyxl

    date_re = re.compile(r"\d{1,2}\.\d{1,2}\.\d{2,4}")
    wb = openpyxl.load_workbook(path, read_only=True, data_only=True)
    try:
        candidates: list[tuple[date, int, str]] = []
        for idx, name in enumerate(wb.sheetnames):
            if not name:
                continue
            compact = name.replace(" ", "")
            if "88" in compact:
                continue
            if "집회" in name or "전도명단" in name:
                continue
            if "주일" in name and "예배" in name and date_re.search(name):
                d = _sheet_title_date(name)
                if d is not None:
                    candidates.append((d, idx, name))
                else:
                    candidates.append((date.min, idx, name))
        if candidates:
            candidates.sort(key=lambda t: (t[0], t[1]))
            return candidates[-1][2]
        for idx, name in enumerate(wb.sheetnames):
            if "주일예배" in name.replace(" ", "") and date_re.search(name):
                return name
    finally:
        wb.close()
    return None


class Command(BaseCommand):
    help = "admin/1234 + 조직도 + (엑셀) 교적·주일 출석까지 초기화"

    def add_arguments(self, parser):
        parser.add_argument(
            "--xlsx",
            type=str,
            default=str(DEFAULT_XLSX),
            help=f"예배 출석 명단 엑셀 (기본: {DEFAULT_XLSX})",
        )
        parser.add_argument(
            "--roster-sheet",
            default="주일 88",
            help="팀·이름 로스터 시트 (기본: 주일 88)",
        )
        parser.add_argument(
            "--sunday-sheet",
            default="",
            help='주일 참석자 명단 시트 (예: "26.03.22 주일예배"). 비우면 파일에서 자동 탐지',
        )
        parser.add_argument(
            "--skip-excel",
            action="store_true",
            help="엑셀 관련 단계 생략 (교적·주일 출석 임포트 안 함)",
        )
        parser.add_argument(
            "--skip-sunday",
            action="store_true",
            help="주일 출석 라인 임포트만 건너뜀 (로스터는 유지)",
        )

    def handle(self, *args, **options):
        self.stdout.write("=== 1/4 bootstrap_admin (admin / 1234) ===")
        call_command("bootstrap_admin", stdout=self.stdout, stderr=self.stderr)

        self.stdout.write(self.style.SUCCESS("=== 2/4 seed_org_chart ==="))
        call_command("seed_org_chart", stdout=self.stdout, stderr=self.stderr)

        admin = get_user_model().objects.filter(username="admin").first()
        if admin:
            rl = RoleLevel.objects.filter(code="pastor").first()
            if rl and admin.role_level_id != rl.id:
                admin.role_level = rl
                admin.save(update_fields=["role_level"])
                self.stdout.write("  admin 계정에 직급 목사(pastor) 연결")

        skip_excel = options["skip_excel"]
        path: Path | None = None if skip_excel else resolve_attendance_workbook(options["xlsx"])

        if skip_excel:
            self.stdout.write(
                self.style.WARNING("=== 3/4 엑셀 스킵 (--skip-excel) ===")
            )
        elif path is None:
            fix = _app_dir() / "fixtures"
            raise CommandError(
                f"엑셀을 찾을 수 없습니다. (`{WORKBOOK_BASENAME}`)\n"
                f"  • 저장소: {fix}/\n"
                "  • 또는 ~/Downloads/\n"
                "  • 또는 환경변수 JCC_SEED_XLSX=/절대/경로.xlsx\n"
                "  • 또는 python manage.py bootstrap_dev --xlsx /경로/파일.xlsx\n"
                f"(기본으로 시도한 첫 경로: {Path(options['xlsx']).expanduser()})"
            )
        else:
            if Path(options["xlsx"]).expanduser().resolve() != path:
                self.stdout.write(
                    self.style.NOTICE(f"엑셀 사용: {path}"),
                )
            self.stdout.write(self.style.SUCCESS(f"=== 3/4 seed_youth_roster ({path.name}) ==="))
            call_command(
                "seed_youth_roster",
                str(path),
                sheet=options["roster_sheet"],
                stdout=self.stdout,
                stderr=self.stderr,
            )

        if skip_excel or path is None:
            self.stdout.write(
                self.style.WARNING("=== 4/4 주일 출석 임포트 스킵 (엑셀 없음) ===")
            )
        elif options["skip_sunday"]:
            self.stdout.write(
                self.style.WARNING("=== 4/4 주일 출석 임포트 스킵 (--skip-sunday) ===")
            )
        else:
            sheet = (options["sunday_sheet"] or "").strip()
            if not sheet:
                sheet = _pick_sunday_attendance_sheet(path) or ""
            if not sheet:
                raise CommandError(
                    "주일 예배 출석 시트를 자동으로 찾지 못했습니다. "
                    "워크북에 '26.03.22 주일예배' 형태 시트가 있는지 확인하거나 "
                    f'`--sunday-sheet "시트이름"` 으로 지정하세요.\n'
                    f"파일: {path}"
                )
            self.stdout.write(
                self.style.SUCCESS(f"=== 4/4 import_sunday_attendance_xlsx (시트: {sheet!r}) ===")
            )
            call_command(
                "import_sunday_attendance_xlsx",
                str(path),
                sheet=sheet,
                division_code="youth",
                create_missing_members=True,
                stdout=self.stdout,
                stderr=self.stderr,
            )

        self.stdout.write(
            self.style.SUCCESS(
                "\n완료. Admin: http://127.0.0.1:8000/admin/  (admin / 1234)"
            )
        )
