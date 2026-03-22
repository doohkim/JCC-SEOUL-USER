"""
청년부 예배 출석 엑셀(부서 회장단 + 팀 열) 파싱.

단일 시트 / 워크북 전체 시트 스캔 지원.
"""

from __future__ import annotations

import hashlib
import re
from collections import defaultdict
from pathlib import Path
from typing import TYPE_CHECKING

if TYPE_CHECKING:
    pass

# 엑셀 팀명 → Team.code (청년부 내 유일)
TEAM_SLUG = {
    "현혜팀": "hyeonhye",
    "주영팀": "juyoung",
    "재영팀": "jaeyoung",
    "주남팀": "junam",
    "광진팀": "gwangjin",
    "미영팀": "miyoung",
    "예린팀": "yerin",
    "연주팀": "yeonju",
    "희원팀": "heewon",
    "경업팀": "gyeongeop",
}


def clean_cell(s):
    if s is None or not isinstance(s, str):
        return None
    s = s.strip()
    if not s:
        return None
    if s in ("회장단", "현장", "부서 회장단", "인천"):
        return None
    if s in ("온", "V") or re.fullmatch(r"[\d.]+", s):
        return None
    s = re.sub(r"\s*(팀장|셀장|부장)$", "", s).strip()
    if len(s) < 2:
        return None
    return s


def parse_sheet(rows: list[tuple]) -> tuple[list[tuple[int, str]], dict[str, set[str]]]:
    """(team_cols, team_members) team_cols: (col_index, team_name)"""
    header_idx = None
    team_cols: list[tuple[int, str]] = []
    for i, row in enumerate(rows):
        if not row:
            continue
        if any(x == "부서 회장단" for x in row if x is not None):
            for j, cell in enumerate(row):
                if isinstance(cell, str) and "팀" in cell and cell != "부서 회장단":
                    team_cols.append((j, cell.replace(" ", "")))
            header_idx = i
            break
    if header_idx is None or not team_cols:
        raise ValueError("시트에서 '부서 회장단' 행 또는 팀 열을 찾지 못했습니다.")

    team_members: dict[str, set[str]] = defaultdict(set)
    for i in range(header_idx + 1, min(header_idx + 120, len(rows))):
        row = rows[i]
        if not row:
            continue
        first = next((c for c in row if c is not None), None)
        if isinstance(first, str) and any(
            k in first for k in ("합계", "참석", "명단", "명 입니다", "총 합")
        ):
            continue
        for col_idx, team_name in team_cols:
            if col_idx >= len(row):
                continue
            nm = clean_cell(row[col_idx])
            if nm:
                team_members[team_name].add(nm)
    return team_cols, team_members


def ascii_username_base(display_name: str) -> str:
    s = "".join(c.lower() for c in display_name if c.isascii() and c.isalnum())
    if s:
        return s[:24]
    return "n" + hashlib.md5(display_name.encode("utf-8")).hexdigest()[:10]


def load_workbook_rows(path: Path, sheet_name: str) -> list[tuple]:
    import openpyxl

    wb = openpyxl.load_workbook(path, read_only=True, data_only=True)
    if sheet_name not in wb.sheetnames:
        avail = ", ".join(wb.sheetnames[:12])
        wb.close()
        raise ValueError(f"시트 '{sheet_name}' 없음. 예: {avail}")
    ws = wb[sheet_name]
    rows = list(ws.iter_rows(values_only=True))
    wb.close()
    return rows


def iter_parseable_sheets(path: Path) -> dict[str, tuple[list[tuple[int, str]], dict[str, set[str]]]]:
    """워크북의 모든 시트를 열어 파싱 가능한 시트만 반환."""
    import openpyxl

    out: dict[str, tuple[list[tuple[int, str]], dict[str, set[str]]]] = {}
    # read_only 는 시트 전환 시 환경에 따라 불안정할 수 있어 일반 모드 사용
    wb = openpyxl.load_workbook(path, read_only=False, data_only=True)
    try:
        for name in wb.sheetnames:
            ws = wb[name]
            rows = list(ws.iter_rows(values_only=True))
            try:
                team_cols, team_members = parse_sheet(rows)
            except ValueError:
                continue
            if team_members:
                out[name] = (team_cols, team_members)
    finally:
        wb.close()
    return out
