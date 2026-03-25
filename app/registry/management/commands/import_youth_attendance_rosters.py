"""
нПілНФ мХДлЮШмЭШ мШИл∞∞ мґЬмДЭ мЧСмЕА(лПЩмЭЉ нПђлІЈ)мЭД мЭљмЦі Member¬Јм≤≠лЕДлґА нМА мЖМмЖНк≥Љ
``WorshipRosterScope`` / ``WorshipRosterEntry`` мЧР л∞ШмШБнХ©лЛИлЛ§.

к≤љл°ЬмЧРмДЬ мЧ∞лПД¬ЈмШИл∞∞ кµђлґД¬ЈлґАл•Љ мґФл°†нХ©лЛИлЛ§.

- **мЭЄм≤Ь**: к≤љл°ЬмЧР 1~4лґАк∞А мЧЖмЬЉл©і **3лґА**л°Ь м†АмЮ•.
- **мДЬмЪЄ**: к≤љл°ЬмЧР ``1лґА``~``4лґА`` (лШРлКФ ``1bu`` лУ±) к∞А мЮИмЦімХЉ нХ©лЛИлЛ§.
- **мШ®лЭЉмЭЄ / мІАкµРнЪМ**: ``session_part`` = 0 (лґА нХілЛє мЧЖмЭМ). мІАкµРнЪМлКФ ``мІАкµРнЪМ/вЧЛвЧЛ/нММмЭЉ.xlsx`` мЭШ вЧЛвЧЛмЭД ``branch_label`` л°Ь мВђмЪ©.

мЛЬнКЄ:

- кЄ∞л≥Є: ``м£ЉмЭЉ 88`` мЛЬнКЄлІМ (мЧЖмЬЉл©і мЫМнБђлґБмЧРмДЬ **м≤Ђ л≤ИмІЄ** нММмЛ± к∞АлК• мЛЬнКЄ 1к∞ЬлІМ).
- ``--single-sheet "мЭіл¶Д"``: нХілЛє мЛЬнКЄлІМ.
- ``--all-sheets``: нММмЛ± к∞АлК•нХЬ **л™®лУ†** мЛЬнКЄл•Љ к∞Бк∞Б л≥ДлПД кµђлґД(``snapshot_label``)мЬЉл°Ь м†АмЮ•.

мВђмЪ© мШИ::

    python manage.py import_youth_attendance_rosters /path/to/rosters
    python manage.py import_youth_attendance_rosters /path/to/rosters --dry-run
    python manage.py import_youth_attendance_rosters /path/to/rosters --division-code youth
    python manage.py import_youth_attendance_rosters /path/to/rosters --all-sheets
"""

from __future__ import annotations

import logging
import re
from pathlib import Path

logger = logging.getLogger("registry.import")

from django.core.exceptions import ValidationError
from django.core.management.base import BaseCommand, CommandError
from django.db import transaction
from attendance.models import WorshipRosterEntry, WorshipRosterScope
from registry.importers.roster_path_context import infer_roster_path_context
from registry.importers.youth_roster_xlsx import (
    TEAM_SLUG,
    ascii_username_base,
    iter_parseable_sheets,
    parse_sheet,
)
from registry.models import Member, MemberDivisionTeam
from users.models import Division, Team


class Command(BaseCommand):
    help = "нПілНФ лВі мґЬмДЭ мЧСмЕАмЭД мЧ∞лПД¬ЈмШИл∞∞кµђлґД¬ЈлґАл≥Дл°Ь мЮДнПђнКЄнХ©лЛИлЛ§."

    def add_arguments(self, parser):
        parser.add_argument(
            "root",
            type=str,
            help="мЧСмЕАлУ§мЭі лУ§мЦі мЮИлКФ л£®нКЄ нПілНФ",
        )
        parser.add_argument(
            "--division-code",
            default="youth",
            help="м≤≠лЕДлґА Division code (кЄ∞л≥Є: youth)",
        )
        parser.add_argument(
            "--dry-run",
            action="store_true",
            help="DBмЧР мУ∞мІА мХКк≥† нММмЛ±¬ЈмґФл°† к≤∞к≥ЉлІМ мґЬл†•",
        )
        parser.add_argument(
            "--preferred-sheet",
            default="м£ЉмЭЉ 88",
            help="кЄ∞л≥Є л™®лУЬмЧРмДЬ мЪ∞мД† мВђмЪ©нХ† мЛЬнКЄл™Е (мЧЖмЬЉл©і м≤Ђ нММмЛ± к∞АлК• мЛЬнКЄ)",
        )
        parser.add_argument(
            "--single-sheet",
            type=str,
            default="",
            help="мІАм†Х мЛЬ мЭі мЛЬнКЄлІМ мЭљмЭМ (--preferred-sheet¬Ј--all-sheetsл≥ілЛ§ мЪ∞мД†)",
        )
        parser.add_argument(
            "--all-sheets",
            action="store_true",
            help="нММмЛ± к∞АлК•нХЬ л™®лУ† мЛЬнКЄл•Љ мЛЬнКЄл™ЕмЬЉл°Ь snapshot_label кµђлґДнХШмЧђ м†АмЮ•",
        )

    def handle(self, *args, **options):
        root = Path(options["root"]).expanduser().resolve()
        division_code = options["division_code"]
        dry = options["dry_run"]
        single_sheet = (options["single_sheet"] or "").strip()
        preferred_sheet = (options["preferred_sheet"] or "").strip() or "м£ЉмЭЉ 88"
        all_sheets_mode = bool(options["all_sheets"]) and not single_sheet

        if not root.is_dir():
            raise CommandError(f"нПілНФк∞А мХДлЛЩлЛИлЛ§: {root}")

        try:
            import openpyxl
        except ImportError as e:
            raise CommandError(
                "openpyxlмЭі нХДмЪФнХ©лЛИлЛ§. poetry install лШРлКФ pip install openpyxl"
            ) from e

        xlsx_files = sorted(root.rglob("*.xlsx"))
        # мЮДмЛЬ/мИ®кєА м†ЬмЩЄ
        xlsx_files = [
            p
            for p in xlsx_files
            if not p.name.startswith("~$") and "/." not in str(p)
        ]

        if not xlsx_files:
            raise CommandError(f"xlsx нММмЭЉмЭі мЧЖмКµлЛИлЛ§: {root}")

        self.stdout.write(f"лМАмГБ нММмЭЉ {len(xlsx_files)}к∞Ь (л£®нКЄ: {root})")

        parsed_files: list[tuple[Path, object, dict]] = []
        skipped = 0

        for path in xlsx_files:
            rel = path.relative_to(root)
            ctx = infer_roster_path_context(rel)
            if ctx is None:
                self.stdout.write(
                    self.style.WARNING(f"к≤љл°Ь мґФл°† мЛ§нМ®(мК§нВµ): {rel}")
                )
                skipped += 1
                continue

            if single_sheet:
                wb = openpyxl.load_workbook(path, read_only=True, data_only=True)
                if single_sheet not in wb.sheetnames:
                    wb.close()
                    self.stdout.write(
                        self.style.WARNING(
                            f"мЛЬнКЄ мЧЖмЭМ(мК§нВµ): {rel} вЖТ '{single_sheet}'"
                        )
                    )
                    skipped += 1
                    continue
                ws = wb[single_sheet]
                rows = list(ws.iter_rows(values_only=True))
                wb.close()
                try:
                    team_cols, team_members = parse_sheet(rows)
                except ValueError as e:
                    self.stdout.write(
                        self.style.WARNING(f"нММмЛ± мЛ§нМ®(мК§нВµ): {rel} вАФ {e}")
                    )
                    skipped += 1
                    continue
                sheets_data = {single_sheet: (team_cols, team_members)}
            else:
                sheets_data = iter_parseable_sheets(path)
                if not sheets_data:
                    self.stdout.write(
                        self.style.WARNING(
                            f"мЭљмЭД мЛЬнКЄ мЧЖмЭМ(мК§нВµ): {rel} (лґАмДЬ нЪМмЮ•лЛ® нПђлІЈ?)"
                        )
                    )
                    skipped += 1
                    continue
                if not all_sheets_mode:
                    if preferred_sheet in sheets_data:
                        sheets_data = {preferred_sheet: sheets_data[preferred_sheet]}
                    else:
                        first_sn = next(iter(sheets_data))
                        sheets_data = {first_sn: sheets_data[first_sn]}
                        self.stdout.write(
                            self.style.WARNING(
                                f"'{preferred_sheet}' мЧЖмЭМ вЖТ '{first_sn}' мВђмЪ©: {rel}"
                            )
                        )

            parsed_files.append((path, ctx, sheets_data))

        if dry:
            for path, ctx, sheets_data in parsed_files:
                rel = path.relative_to(root)
                mode = "all-sheets" if all_sheets_mode else "лЛ®мЭЉ мЛЬнКЄ"
                self.stdout.write(
                    f"[dry] {rel} вЖТ {ctx.year} {ctx.venue} лґА={ctx.session_part} "
                    f"мІАкµРнЪМ={ctx.branch_label!r} ({mode}) мЛЬнКЄ={list(sheets_data.keys())}"
                )
                for sn, (_, tm) in sheets_data.items():
                    snap = sn[:200] if all_sheets_mode else ""
                    n = sum(len(v) for v in tm.values())
                    self.stdout.write(
                        f"        ¬Ј {sn} snapshot={snap!r}: нМА {len(tm)}к∞Ь, мЭіл¶Д {n}к±і"
                    )
            self.stdout.write(
                self.style.WARNING(
                    f"dry-run мЩДл£М (мК§нВµ {skipped}, нММмЛ± мД±к≥µ {len(parsed_files)})"
                )
            )
            return

        with transaction.atomic():
            div, _ = Division.objects.get_or_create(
                code=division_code,
                defaults={"name": "м≤≠лЕДлґА", "sort_order": 10},
            )

            used_ik: set[str] = set(
                Member.objects.exclude(import_key="").values_list("import_key", flat=True)
            )
            by_name: dict[str, Member] = {}
            for m in Member.objects.all():
                key = re.sub(r"\s+", "", (m.name or "").strip())
                if key:
                    by_name[key] = m

            created_members = 0
            mdt_created = 0
            scope_created = 0
            entry_created = 0
            entry_updated = 0
            entry_skipped_conflict = 0

            for path, ctx, sheets_data in parsed_files:
                rel_str = str(path.relative_to(root))
                for sheet_name, (team_cols, team_members) in sheets_data.items():
                    snapshot_label = sheet_name[:200] if all_sheets_mode else ""
                    scope, sc = WorshipRosterScope.objects.get_or_create(
                        division=div,
                        venue=ctx.venue,
                        year=ctx.year,
                        session_part=ctx.session_part,
                        branch_label=ctx.branch_label or "",
                        snapshot_label=snapshot_label,
                        defaults={"sort_order": 0},
                    )
                    if sc:
                        scope_created += 1

                    team_objs: dict[str, Team] = {}
                    for order, (_, raw_name) in enumerate(team_cols):
                        slug = TEAM_SLUG.get(raw_name)
                        if not slug:
                            self.stdout.write(
                                self.style.WARNING(
                                    f"мХМ мИШ мЧЖлКФ нМА мК§нВµ: {raw_name} ({rel_str})"
                                )
                            )
                            continue
                        t, _ = Team.objects.get_or_create(
                            division=div,
                            code=slug,
                            defaults={"name": raw_name, "sort_order": order},
                        )
                        if t.name != raw_name:
                            t.name = raw_name
                            t.sort_order = order
                            t.save(update_fields=["name", "sort_order"])
                        team_objs[raw_name] = t

                    for team_name, names in sorted(team_members.items()):
                        team = team_objs.get(team_name)
                        if not team:
                            continue
                        for display in sorted(names):
                            key = re.sub(r"\s+", "", display)
                            member = by_name.get(key)
                            if member is None:
                                base = (ascii_username_base(display) or "member")[:64]
                                ik = base
                                n = 1
                                while ik in used_ik:
                                    ik = f"{base}_{n}"[:64]
                                    n += 1
                                used_ik.add(ik)
                                member = Member.objects.create(
                                    name=display[:50],
                                    import_key=ik,
                                )
                                by_name[key] = member
                                created_members += 1

                            mdt, mcreated = MemberDivisionTeam.objects.get_or_create(
                                member=member,
                                division=div,
                                defaults={
                                    "team": team,
                                    "is_primary": not member.division_teams.filter(
                                        division=div, is_primary=True
                                    ).exists(),
                                    "sort_order": 0,
                                },
                            )
                            if mcreated:
                                mdt_created += 1
                            else:
                                if mdt.team_id != team.id:
                                    mdt.team = team
                                    mdt.save(update_fields=["team"])
                                if not member.division_teams.filter(
                                    division=div, is_primary=True
                                ).exists():
                                    mdt.is_primary = True
                                    mdt.save(update_fields=["is_primary"])

                            try:
                                entry = WorshipRosterEntry.objects.get(
                                    scope=scope, member=member
                                )
                            except WorshipRosterEntry.DoesNotExist:
                                entry = WorshipRosterEntry(
                                    scope=scope,
                                    member=member,
                                    team=team,
                                    source_rel_path=rel_str,
                                    sheet_name=sheet_name,
                                )
                                try:
                                    entry.full_clean()
                                    entry.save()
                                except ValidationError as e:
                                    entry_skipped_conflict += 1
                                    msg = (
                                        f"л™ЕлЛ® к≤Ам¶Э мЛ§нМ®(м†АмЮ• мХИ нХ®): {member.name} "
                                        f"scope={scope} {rel_str} вАФ "
                                        f"{e.message_dict or e.messages}"
                                    )
                                    self.stdout.write(self.style.WARNING(msg))
                                    logger.warning(msg)
                                    continue
                                entry_created += 1
                            else:
                                changed = False
                                if entry.team_id != team.id:
                                    entry.team = team
                                    changed = True
                                if entry.source_rel_path != rel_str:
                                    entry.source_rel_path = rel_str
                                    changed = True
                                if entry.sheet_name != sheet_name:
                                    entry.sheet_name = sheet_name
                                    changed = True
                                if changed:
                                    try:
                                        entry.full_clean()
                                    except ValidationError as e:
                                        entry_skipped_conflict += 1
                                        msg = (
                                            f"л™ЕлЛ® к∞±мЛ† мК§нВµ(к≤Ам¶Э мЛ§нМ®): {member.name} "
                                            f"scope={scope} вАФ {e.message_dict or e.messages}"
                                        )
                                        self.stdout.write(self.style.WARNING(msg))
                                        logger.warning(msg)
                                        continue
                                    entry.save()
                                    entry_updated += 1

        self.stdout.write(
            self.style.SUCCESS(
                f"мЩДл£М: мК§нВµ {skipped}, кµђлґД мЛ†кЈЬ {scope_created}, "
                f"Member мЛ†кЈЬ {created_members}, мЖМмЖН мЛ†кЈЬ {mdt_created}, "
                f"л™ЕлЛ®нЦЙ мЛ†кЈЬ {entry_created}, л™ЕлЛ®нЦЙ к∞±мЛ† {entry_updated}, "
                f"л™ЕлЛ® к≤Ам¶Э мК§нВµ {entry_skipped_conflict}"
            )
        )
